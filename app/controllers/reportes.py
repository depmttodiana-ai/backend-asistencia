from fastapi import APIRouter, Depends, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.core.database import get_db
from app.core.auth_deps import coordinador_or_admin
from app.controllers.asistencia import generar_reporte_nomina_comun
from app.models.models_empleados import Empleado as EmpleadoReal
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from datetime import date

router = APIRouter()

# --- CONSTANTES DE ESTILO ---
ANCHO_1CM = 4.71
COLOR_HEADER_AZUL = "0070C0"
COLOR_HEADER_GRIS = "E7E6E6"
COLOR_HEADER_AMARILLO = "FFCC00"
COLOR_ROJO = "FF0000"
CARGOS_JEFES = [
    "DIRECTOR AREA DE MANTENIMIENTO",
    "DIRECTOR DE MANTENIMIENTO",
    "SUPERVISOR DE MANTENIMIENTO",
    "COORDINADOR",
]


def aplicar_estilos_base(
    ws, logo_bytes, fecha_desde, fecha_hasta, titulo="REPORTE DE ASISTENCIAS"
):
    """Configura el layout inicial, logo y título global"""
    ws.sheet_view.showGridlines = False

    # Anchos
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 10
    for i in range(3, 26):
        ws.column_dimensions[get_column_letter(i)].width = ANCHO_1CM

    # Alturas para logo
    for r in range(1, 8):
        ws.row_dimensions[r].height = 22

    # Logo
    if logo_bytes:
        try:
            img = ExcelImage(io.BytesIO(logo_bytes))
            img.width = 1202
            img.height = 202
            ws.add_image(img, "A1")
            pass
        except Exception:
            pass

    # Títulos
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A8:Y8")
    ws["A8"] = titulo
    ws["A8"].font = Font(bold=True, size=16)
    ws["A8"].alignment = align_center

    ws.merge_cells("A9:Y9")
    ws["A9"] = (
        f"SEMANA DEL {fecha_desde.strftime('%d/%m/%Y')} AL {fecha_hasta.strftime('%d/%m/%Y')}"
    )
    ws["A9"].fill = PatternFill(
        start_color=COLOR_HEADER_AZUL, end_color=COLOR_HEADER_AZUL, fill_type="solid"
    )
    ws["A9"].font = Font(color="FFFFFF", bold=True)
    ws["A9"].alignment = align_center


def dibujar_encabezados_tabla(ws, fila_inicial):
    """Dibuja los encabezados de columnas (Asistencia, HE, etc.)"""
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    fill_gris = PatternFill(
        start_color=COLOR_HEADER_GRIS, end_color=COLOR_HEADER_GRIS, fill_type="solid"
    )

    fila = fila_inicial

    # Super-encabezados (Asistencia, HE Diurnas, HE Nocturnas)
    grupos = [
        ("C", "I", "ASISTENCIA"),
        ("J", "P", "H.E. DIURNAS"),
        ("Q", "W", "H.E. NOCTURNAS"),
    ]
    for ini, fin, txt in grupos:
        ws.merge_cells(f"{ini}{fila}:{fin}{fila}")
        c = ws[f"{ini}{fila}"]
        c.value = txt
        c.alignment = align_center
        c.font = Font(bold=True, size=7)
        c.fill = fill_gris
        c.border = thin_border

    # Sub-encabezados (Nombres, dias...)
    fila += 1
    headers = (
        ["Nombre del Trabajador", "Código"]
        + (["L", "M", "M", "J", "V", "S", "D"] * 3)
        + ["FE", "Σ H.E."]
    )

    for col, val in enumerate(headers, 1):
        c = ws.cell(row=fila, column=col, value=val)
        c.font = Font(bold=True, size=8)
        c.alignment = align_center
        c.border = thin_border
        c.fill = fill_gris

    return fila + 1


def dibujar_encabezados_resumen(ws, fila_inicial):
    """
    Dibuja los encabezados expandidos para el reporte de SOLO HORAS EXTRAS (Estilo Diario).
    Columnas: Nombre, Código, [L..D Diurnas], [L..D Nocturnas], FE, Total
    """
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    fill_gris = PatternFill(
        start_color=COLOR_HEADER_GRIS, end_color=COLOR_HEADER_GRIS, fill_type="solid"
    )

    fila = fila_inicial

    # Super-Encabezados
    grupos = [
        ("C", "I", "H.E. DIURNAS"),
        ("J", "P", "H.E. NOCTURNAS"),
    ]
    for ini, fin, txt in grupos:
        ws.merge_cells(f"{ini}{fila}:{fin}{fila}")
        c = ws[f"{ini}{fila}"]
        c.value = txt
        c.alignment = align_center
        c.font = Font(bold=True, size=7)
        c.fill = fill_gris
        c.border = thin_border

    # Headers Dias
    fila += 1
    headers = (
        ["Nombre del Trabajador", "Código"]
        + (["L", "M", "M", "J", "V", "S", "D"] * 2)
        + ["FE", "TOTAL"]
    )

    for col, val in enumerate(headers, 1):
        c = ws.cell(row=fila, column=col, value=val)
        c.font = Font(bold=True, size=8)
        c.alignment = align_center
        c.border = thin_border
        c.fill = fill_gris

    return fila + 1


def escribir_fila_empleado(ws, fila, emp_data):
    """Escribe los datos de un empleado (Reporte Completo)"""
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Datos fijos
    c_nom = ws.cell(row=fila, column=1, value=emp_data["Nombre"])
    c_nom.border = thin_border
    c_cod = ws.cell(row=fila, column=2, value=emp_data["Codigo"])
    c_cod.border = thin_border
    c_cod.alignment = align_center

    # Datos dinámicos
    values = (
        emp_data["Asistencia"]
        + emp_data["HE_Diurna"]
        + emp_data["HE_Nocturna"]
        + [emp_data["Feriado"], emp_data["Total_HE"]]
    )

    for i, val in enumerate(values):
        col = 3 + i
        if col == 25:
            cell = ws.cell(
                row=fila,
                column=col,
                value=f'=IF(SUM(J{fila}:W{fila})>0, SUM(J{fila}:W{fila}), "")',
            )
        else:
            cell = ws.cell(row=fila, column=col, value=val)

        cell.border = thin_border
        cell.alignment = align_center
        cell.font = Font(size=8)

        if 3 <= col <= 9:  # Asistencias
            val_str = str(val).upper().strip()
            if val_str not in ["X", ""]:
                cell.font = Font(size=8, color=COLOR_ROJO, bold=True)
        elif col == 24:  # Feriados
            if isinstance(val, (int, float)) and val > 0:
                cell.font = Font(size=8, color=COLOR_ROJO, bold=True)


def escribir_fila_resumen_he(ws, fila, emp_data):
    """Escribe los datos de un empleado (Reporte Resumido HE - Expandido)"""
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 1. Nombre y Código
    c_nom = ws.cell(row=fila, column=1, value=emp_data["Nombre"])
    c_nom.border = thin_border
    c_cod = ws.cell(row=fila, column=2, value=emp_data["Codigo"])
    c_cod.border = thin_border
    c_cod.alignment = align_center

    # 2. Datos Dinámicos (HE D + HE N + FE + Total)
    values = emp_data["HE_Diurna"] + emp_data["HE_Nocturna"] + [emp_data["Feriado"], ""]

    for i, val in enumerate(values):
        col = 3 + i
        if col == 18:
            cell = ws.cell(
                row=fila,
                column=col,
                value=f'=IF(SUM(C{fila}:P{fila})>0, SUM(C{fila}:P{fila}), "")',
            )
        else:
            cell = ws.cell(row=fila, column=col, value=val)

        cell.border = thin_border
        cell.alignment = align_center
        cell.font = Font(size=8)

        # Color rojo para feriados si > 0
        if col == 17 and isinstance(val, (int, float)) and val > 0:
            cell.font = Font(size=8, color=COLOR_ROJO, bold=True)


def dibujar_seccion_turno(ws, fila_inicial, titulo_turno, lista_empleados):
    """Dibuja un bloque completo (Reporte Completo)"""
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells(f"A{fila_inicial}:Y{fila_inicial}")
    c = ws[f"A{fila_inicial}"]
    c.value = titulo_turno
    c.font = Font(bold=True, size=11)
    c.alignment = align_center
    c.fill = PatternFill(
        start_color=COLOR_HEADER_AMARILLO,
        end_color=COLOR_HEADER_AMARILLO,
        fill_type="solid",
    )
    c.border = thin_border

    fila = fila_inicial + 1
    fila = dibujar_encabezados_tabla(ws, fila)

    for emp in lista_empleados:
        escribir_fila_empleado(ws, fila, emp)
        fila += 1

    return fila


def dibujar_seccion_resumen_he(ws, fila_inicial, titulo_turno, lista_empleados):
    """Dibuja un bloque completo (Reporte Resumido HE)"""
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Merge hasta R (Col 18) para este reporte expandido
    ws.merge_cells(f"A{fila_inicial}:R{fila_inicial}")
    c = ws[f"A{fila_inicial}"]
    c.value = titulo_turno
    c.font = Font(bold=True, size=11)
    c.alignment = align_center
    c.fill = PatternFill(
        start_color=COLOR_HEADER_AMARILLO,
        end_color=COLOR_HEADER_AMARILLO,
        fill_type="solid",
    )
    c.border = thin_border

    fila = fila_inicial + 1
    fila = dibujar_encabezados_resumen(ws, fila)

    for emp in lista_empleados:
        escribir_fila_resumen_he(ws, fila, emp)
        fila += 1

    return fila


def dibujar_firma_fernelis(ws, fila_actual, ancho_cols="Y"):
    """Dibuja el bloque de firma de Fernelis"""
    ws.merge_cells(f"A{fila_actual}:{ancho_cols}{fila_actual}")
    ws[f"A{fila_actual}"] = "________________________________________________"
    ws[f"A{fila_actual}"].alignment = Alignment(horizontal="center")

    fila_actual += 1
    ws.merge_cells(f"A{fila_actual}:{ancho_cols}{fila_actual}")
    c_firma = ws[f"A{fila_actual}"]
    c_firma.value = "FERNELIS"
    c_firma.font = Font(bold=True, size=11)
    c_firma.alignment = Alignment(horizontal="center")

    return fila_actual + 1


# --- GENERADORES EXCEL (SEPARADOS) ---


def crear_excel_empleados(grupos: dict, fecha_desde, fecha_hasta, logo_bytes):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Empleados"
    aplicar_estilos_base(
        ws,
        logo_bytes,
        fecha_desde,
        fecha_hasta,
        titulo="REPORTE DE ASISTENCIAS - OPERATIVOS",
    )
    fila_actual = 11

    if grupos["Diurno"]:
        fila_actual = dibujar_seccion_turno(
            ws, fila_actual, "TURNO DIURNO", grupos["Diurno"]
        )
        fila_actual += 1
    if grupos["Nocturno"]:
        fila_actual = dibujar_seccion_turno(
            ws, fila_actual, "TURNO NOCTURNO", grupos["Nocturno"]
        )
        fila_actual += 2

    dibujar_firma_fernelis(ws, fila_actual)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def crear_excel_coordinadores(
    grupos: dict, fecha_desde, fecha_hasta, logo_supervisores_bytes
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Coordinadores"
    aplicar_estilos_base(
        ws,
        logo_supervisores_bytes,
        fecha_desde,
        fecha_hasta,
        titulo="REPORTE DE ASISTENCIAS - JEFATURA",
    )
    fila_actual = 11

    if grupos["Jefes"]:
        fila_actual = dibujar_seccion_turno(
            ws, fila_actual, "DIRECCIÓN Y SUPERVISIÓN DE MANTENIMIENTO", grupos["Jefes"]
        )
        fila_actual += 2
        dibujar_firma_fernelis(ws, fila_actual)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def crear_excel_resumen_he_empleados(
    grupos: dict, fecha_desde, fecha_hasta, logo_bytes
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen HE Empleados"

    # Custom headers visual approach for summary
    ws.sheet_view.showGridlines = False

    # Logo y Headers Manuales (Simplificado vs aplicar_estilos_base)
    if logo_bytes:
        try:
            img = ExcelImage(io.BytesIO(logo_bytes))
            img.width = 990
            img.height = 202
            ws.add_image(img, "A1")
        except:
            pass

    for r in range(1, 8):
        ws.row_dimensions[r].height = 22

    # Anchos
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 10
    for i in range(3, 17):  # C hasta P
        ws.column_dimensions[get_column_letter(i)].width = ANCHO_1CM

    fila_actual = 8
    ws.merge_cells(f"A{fila_actual}:R{fila_actual}")
    ws[f"A{fila_actual}"] = "RESUMEN HORAS EXTRAS - OPERATIVOS"
    ws[f"A{fila_actual}"].font = Font(bold=True, size=16)
    ws[f"A{fila_actual}"].alignment = Alignment(horizontal="center", vertical="center")

    fila_actual += 1
    ws.merge_cells(f"A{fila_actual}:R{fila_actual}")
    ws[f"A{fila_actual}"] = (
        f"SEMANA DEL {fecha_desde.strftime('%d/%m/%Y')} AL {fecha_hasta.strftime('%d/%m/%Y')}"
    )
    ws[f"A{fila_actual}"].fill = PatternFill(
        start_color=COLOR_HEADER_AZUL, end_color=COLOR_HEADER_AZUL, fill_type="solid"
    )
    ws[f"A{fila_actual}"].font = Font(color="FFFFFF", bold=True)
    ws[f"A{fila_actual}"].alignment = Alignment(horizontal="center", vertical="center")

    fila_actual += 2

    if grupos["Diurno"]:
        fila_actual = dibujar_seccion_resumen_he(
            ws, fila_actual, "TURNO DIURNO", grupos["Diurno"]
        )
        fila_actual += 1
    if grupos["Nocturno"]:
        fila_actual = dibujar_seccion_resumen_he(
            ws, fila_actual, "TURNO NOCTURNO", grupos["Nocturno"]
        )
        fila_actual += 2

    dibujar_firma_fernelis(ws, fila_actual, ancho_cols="R")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def crear_excel_resumen_he_jefes(grupos: dict, fecha_desde, fecha_hasta, logo_bytes):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen HE Jefes"

    if logo_bytes:
        try:
            img = ExcelImage(io.BytesIO(logo_bytes))
            img.width = 990
            img.height = 202
            ws.add_image(img, "A1")
        except:
            pass

    for r in range(1, 8):
        ws.row_dimensions[r].height = 22

    # Anchos
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 10
    for i in range(3, 17):  # C hasta P
        ws.column_dimensions[get_column_letter(i)].width = ANCHO_1CM

    fila_actual = 8
    ws.merge_cells(f"A{fila_actual}:R{fila_actual}")
    ws[f"A{fila_actual}"] = "RESUMEN HORAS EXTRAS - JEFATURA"
    ws[f"A{fila_actual}"].font = Font(bold=True, size=16)
    ws[f"A{fila_actual}"].alignment = Alignment(horizontal="center", vertical="center")

    fila_actual += 1
    ws.merge_cells(f"A{fila_actual}:R{fila_actual}")
    ws[f"A{fila_actual}"] = (
        f"SEMANA DEL {fecha_desde.strftime('%d/%m/%Y')} AL {fecha_hasta.strftime('%d/%m/%Y')}"
    )
    ws[f"A{fila_actual}"].fill = PatternFill(
        start_color=COLOR_HEADER_AZUL, end_color=COLOR_HEADER_AZUL, fill_type="solid"
    )
    ws[f"A{fila_actual}"].font = Font(color="FFFFFF", bold=True)
    ws[f"A{fila_actual}"].alignment = Alignment(horizontal="center", vertical="center")

    fila_actual += 2

    if grupos["Jefes"]:
        fila_actual = dibujar_seccion_resumen_he(
            ws, fila_actual, "DIRECCIÓN Y SUPERVISIÓN DE MANTENIMIENTO", grupos["Jefes"]
        )
        fila_actual += 2
        dibujar_firma_fernelis(ws, fila_actual, ancho_cols="R")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# --- ENDPOINTS ---


@router.post(
    "/reporte-excel/empleados",
    summary="Descargar Excel Nómina (Solo Empleados)",
    dependencies=[Depends(coordinador_or_admin)],
)
async def descargar_reporte_empleados(
    fecha_desde: date = Form(..., description="Fecha Inicio (YYYY-MM-DD)"),
    fecha_hasta: date = Form(..., description="Fecha Fin (YYYY-MM-DD)"),
    logo: UploadFile = File(None),
    db: Session = Depends(get_db),
):
    todos_empleados = db.query(EmpleadoReal).order_by(EmpleadoReal.apellido).all()
    reporte_data = generar_reporte_nomina_comun(
        fecha_desde, fecha_hasta, todos_empleados, db
    )
    logo_bytes = await logo.read() if logo else None
    grupos = clasificar_empleados(reporte_data)
    excel_file = crear_excel_empleados(grupos, fecha_desde, fecha_hasta, logo_bytes)
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="Reporte_Empleados_{fecha_desde}.xlsx"'
        },
    )


@router.post(
    "/reporte-excel/jefes",
    summary="Descargar Excel Nómina (Solo Jefes)",
    dependencies=[Depends(coordinador_or_admin)],
)
async def descargar_reporte_jefes(
    fecha_desde: date = Form(..., description="Fecha Inicio (YYYY-MM-DD)"),
    fecha_hasta: date = Form(..., description="Fecha Fin (YYYY-MM-DD)"),
    logo: UploadFile = File(None),
    db: Session = Depends(get_db),
):
    todos_empleados = db.query(EmpleadoReal).order_by(EmpleadoReal.apellido).all()
    reporte_data = generar_reporte_nomina_comun(
        fecha_desde, fecha_hasta, todos_empleados, db
    )
    logo_bytes = await logo.read() if logo else None
    grupos = clasificar_empleados(reporte_data)
    excel_file = crear_excel_coordinadores(grupos, fecha_desde, fecha_hasta, logo_bytes)
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="Reporte_Jefes_{fecha_desde}.xlsx"'
        },
    )


@router.post(
    "/reporte-excel/resumen-he/empleados",
    summary="Resumen Horas Extras (Empleados)",
    dependencies=[Depends(coordinador_or_admin)],
)
async def descargar_resumen_he_empleados(
    fecha_desde: date = Form(..., description="Fecha Inicio (YYYY-MM-DD)"),
    fecha_hasta: date = Form(..., description="Fecha Fin (YYYY-MM-DD)"),
    logo: UploadFile = File(None),
    db: Session = Depends(get_db),
):
    """
    Genera Reporte Resumido de Horas Extras para Empleados.
    Columnas: Nombre, Codigo, HE Diurna, HE Nocturna, Feriados, Total.
    """
    todos_empleados = db.query(EmpleadoReal).order_by(EmpleadoReal.apellido).all()
    reporte_data = generar_reporte_nomina_comun(
        fecha_desde, fecha_hasta, todos_empleados, db
    )
    logo_bytes = await logo.read() if logo else None
    grupos = clasificar_empleados(reporte_data)

    excel_file = crear_excel_resumen_he_empleados(
        grupos, fecha_desde, fecha_hasta, logo_bytes
    )

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="Resumen_HE_Empleados_{fecha_desde}.xlsx"'
        },
    )


@router.post(
    "/reporte-excel/resumen-he/jefes",
    summary="Resumen Horas Extras (Jefes)",
    dependencies=[Depends(coordinador_or_admin)],
)
async def descargar_resumen_he_jefes(
    fecha_desde: date = Form(..., description="Fecha Inicio (YYYY-MM-DD)"),
    fecha_hasta: date = Form(..., description="Fecha Fin (YYYY-MM-DD)"),
    logo: UploadFile = File(None),
    db: Session = Depends(get_db),
):
    """
    Genera Reporte Resumido de Horas Extras para Jefes.
    Columnas: Nombre, Codigo, HE Diurna, HE Nocturna, Feriados, Total.
    """
    todos_empleados = db.query(EmpleadoReal).order_by(EmpleadoReal.apellido).all()
    reporte_data = generar_reporte_nomina_comun(
        fecha_desde, fecha_hasta, todos_empleados, db
    )
    logo_bytes = await logo.read() if logo else None
    grupos = clasificar_empleados(reporte_data)

    excel_file = crear_excel_resumen_he_jefes(
        grupos, fecha_desde, fecha_hasta, logo_bytes
    )

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="Resumen_HE_Jefes_{fecha_desde}.xlsx"'
        },
    )


def clasificar_empleados(reporte_data):
    """Helper para separar empleados en listas según cargo/turno"""
    grupos = {"Jefes": [], "Diurno": [], "Nocturno": []}

    # Helper para formateo entero
    def fmt_int(num):
        if isinstance(num, (int, float)) and num > 0:
            return int(round(num))
        return ""

    for reporte_emp in reporte_data.data:
        # Preparar data plana
        lista_asist = []
        lista_hed = []
        lista_hen = []

        detalles = reporte_emp.detalles[:7]

        for det in detalles:
            # Mapeo Visual
            est = det.estado.upper()
            txt = est[:3]
            if "PRESENTE" in est:
                txt = "X"
            elif "FALTA" in est:
                txt = "PVC"
            elif "VACACIONES" in est:
                txt = "V"
            elif "FERIADO" in est:
                txt = "F"
            elif "DESCANSO" in est:
                txt = ""

            lista_asist.append(txt)
            lista_hed.append(fmt_int(det.horas_extras_diurnas))
            lista_hen.append(fmt_int(det.horas_extras_nocturnas))

        # Rellenar
        while len(lista_asist) < 7:
            lista_asist.append("")
            lista_hed.append("")
            lista_hen.append("")

        emp_dict = {
            "Nombre": f"{reporte_emp.nombre_completo}",
            "Codigo": reporte_emp.codigo,
            "Asistencia": lista_asist,
            "HE_Diurna": lista_hed,
            "HE_Nocturna": lista_hen,
            "Feriado": fmt_int(reporte_emp.total_feriados_trabajados),
            "Total_HE": fmt_int(reporte_emp.total_horas_extras_global),
        }

        # LÓGICA DE CLASIFICACIÓN
        cargo_up = (reporte_emp.cargo or "").upper()
        es_jefe = any(c in cargo_up for c in CARGOS_JEFES)

        if es_jefe:
            grupos["Jefes"].append(emp_dict)
        else:
            total_noc = reporte_emp.total_horas_extras_nocturnas
            total_diu = reporte_emp.total_horas_extras_diurnas
            if total_noc > 0 and total_noc >= total_diu:
                grupos["Nocturno"].append(emp_dict)
            else:
                grupos["Diurno"].append(emp_dict)

    return grupos
